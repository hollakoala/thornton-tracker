#!/usr/bin/env python3
"""
Daily Thornton Place / Creekside price tracker.

Ported from Jonas's Cowork routine (thornton-track-cloud) to run as a
GitHub Actions job with no cloud storage dependency - everything lives
in this repo. Reads/writes:

  data/Thornton_Tracker.xlsx   canonical workbook (must already exist -
                                seed it with your current tracker before
                                the first run; this script never creates
                                one from scratch, to protect price history)
  data/price-history.csv       flat, git-diffable append log (one row per
                                unit per run where the price is known)
  docs/index.html              self-contained dashboard, served by GitHub Pages

Constraints preserved from the original routine:
  - Creekside = any unit NOT in Building 2 or 3 (defined by exclusion)
  - Uses TOTAL monthly price, not base rent
  - Never overwrites spec columns (beds/bath/sqft/layout) - flags mismatches only
  - A failed/empty scrape is treated as a no-op, never as "everything off market"
"""

import csv
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "Thornton_Tracker.xlsx"
CSV_PATH = ROOT / "data" / "price-history.csv"
HTML_PATH = ROOT / "docs" / "index.html"

SITE_URL = "https://thornton-place.com/floorplans/"


def scrape_units():
    """Headless-Chromium scrape of the SightMap floor widget. Returns a
    list of dicts, or [] if scraping failed / found nothing (caller must
    treat that as a no-op, not as 'everything went off market')."""
    from playwright.sync_api import sync_playwright

    js = r"""
    (async () => {
      const sleep = ms => new Promise(r => setTimeout(r, ms));
      const all = new Map();
      const floorItems = Array.from(
        document.querySelectorAll('.jd-fp-map-embed__floors-item.swiper-slide')
      );
      const floorsWithAvail = floorItems.filter(el => el.className.includes('--has-avail'));
      const extract = () => {
        document.querySelectorAll('.jd-fp-unit-card.jd-fp-unit-card--row').forEach(c => {
          const txt = c.textContent.trim().replace(/\s+/g, ' ');
          const unitM  = txt.match(/#(\d-\d{3}[A-Z]?)/);
          const totalM = txt.match(/\$([\d,]+\.\d{2})\s*\/mo/);
          const baseM  = txt.match(/\$([\d,]+)\s*Base Rent/);
          const availM = txt.match(/Available\s+(Now|[A-Z][a-z]+\s+\d+)/);
          const bedsM  = txt.match(/(Studio|\d+)\s*bed/i);
          const bathM  = txt.match(/([\d.]+)\s*bath/i);
          const sqftM  = txt.match(/([\d,]+)\s*sq\.?\s*ft/i);
          const layoutM = txt.match(/^([A-Z]\d+[A-Z]?)\s/);
          if (unitM && !all.has(unitM[1])) {
            all.set(unitM[1], { unit: unitM[1], total: totalM?.[1], base: baseM?.[1],
              avail: availM?.[1], beds: bedsM?.[1], bath: bathM?.[1],
              sqft: sqftM?.[1], layout: layoutM?.[1], raw: txt });
          }
        });
      };
      extract();
      for (const item of floorsWithAvail) { item.click(); await sleep(1200); extract(); }
      return Array.from(all.values());
    })()
    """

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(SITE_URL, wait_until="networkidle", timeout=45000)
            page.wait_for_timeout(2500)
            units = page.evaluate(js)
            browser.close()
    except Exception as e:
        print(f"::warning::Scrape failed: {e}")
        return []

    # Creekside = NOT building 2 or 3
    creekside = [u for u in units if u.get("unit") and not u["unit"].startswith(("2-", "3-"))]
    return creekside


def load_tracker():
    if not XLSX_PATH.exists():
        print(f"::error::{XLSX_PATH} not found. Upload your existing Thornton Tracker.xlsx "
              f"into data/ before the first run - this script will not create one from scratch.")
        sys.exit(1)
    return load_workbook(XLSX_PATH)


def find_active_rows(ws):
    """Overview sheet: A=Unit#, B=Creekside?, ... row with A='— Off Market —'
    separates active rows from the off-market archive."""
    active_rows = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        if isinstance(val, str) and "Off Market" in val:
            break
        active_rows.append(row)
    return active_rows


def last_recorded_price(ws, row, last_date_col):
    """Scan newest -> left through the date columns for the most recent numeric price."""
    for col in range(last_date_col, 8, -1):  # column I=9 is the Feb baseline; date cols start at J=10
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            return v, col
    return None, None


def update_tracker_and_history(wb, scraped):
    ws = wb.active  # Overview sheet
    today = datetime.now()
    active_rows = find_active_rows(ws)

    # Find the current last date column (rightmost column before "$ Change")
    header_row = 1
    last_date_col = None
    change_col = None
    for col in range(ws.max_column, 0, -1):
        header = ws.cell(row=header_row, column=col).value
        if header == "$ Change":
            change_col = col
        if isinstance(header, datetime) and last_date_col is None:
            last_date_col = col
    if last_date_col is None or change_col is None:
        print("::error::Could not locate date columns / '$ Change' header - tracker layout unexpected.")
        sys.exit(1)

    tracker_by_unit = {}
    for row in active_rows:
        unit = ws.cell(row=row, column=1).value
        if unit:
            tracker_by_unit[str(unit).lstrip("#")] = row

    scraped_by_unit = {u["unit"]: u for u in scraped}

    price_changes, new_listings, off_market, spec_flags, unchanged = [], [], [], [], []

    # Insert a new date column before "$ Change"
    new_col = change_col
    ws.insert_cols(new_col)
    change_col += 1
    # copy style from previous date column header
    src_header_cell = ws.cell(row=header_row, column=last_date_col)
    new_header_cell = ws.cell(row=header_row, column=new_col)
    new_header_cell.value = today
    new_header_cell.number_format = src_header_cell.number_format
    if src_header_cell.has_style:
        new_header_cell.font = src_header_cell.font.copy()
        new_header_cell.fill = src_header_cell.fill.copy()
        new_header_cell.alignment = src_header_cell.alignment.copy()
        new_header_cell.border = src_header_cell.border.copy()

    csv_rows = []

    for unit, row in tracker_by_unit.items():
        last_price, _ = last_recorded_price(ws, row, last_date_col)
        site = scraped_by_unit.get(unit)

        if site is None:
            ws.cell(row=row, column=8).value = "Off Market"  # Notes column H
            off_market.append((unit, last_price))
            continue

        try:
            current = float(site["total"].replace(",", "")) if site.get("total") else None
        except (ValueError, AttributeError):
            current = None

        if current is not None:
            if last_price is None or abs(current - last_price) > 0.005:
                ws.cell(row=row, column=new_col).value = current
                price_changes.append((unit, last_price, current))
            else:
                unchanged.append(unit)
            csv_rows.append([today.strftime("%Y-%m-%d"), unit, current])

        # spec sanity check only - never auto-write D/E/F
        beds = ws.cell(row=row, column=4).value
        sqft = ws.cell(row=row, column=5).value
        if site.get("sqft") and sqft and str(site["sqft"]).replace(",", "") != str(sqft).replace(",", ""):
            spec_flags.append((unit, "sqft", sqft, site["sqft"]))

    # New listings not already in tracker
    for unit, site in scraped_by_unit.items():
        if unit in tracker_by_unit:
            continue
        new_row = max(active_rows) + 1 if active_rows else 2
        ws.insert_rows(new_row)
        ws.cell(row=new_row, column=1).value = f"#{unit}"
        ws.cell(row=new_row, column=2).value = "\u2705"
        ws.cell(row=new_row, column=4).value = site.get("beds", "")
        ws.cell(row=new_row, column=5).value = site.get("sqft", "")
        ws.cell(row=new_row, column=6).value = site.get("layout", "")
        ws.cell(row=new_row, column=7).value = site.get("avail", "")
        try:
            current = float(site["total"].replace(",", "")) if site.get("total") else None
        except (ValueError, AttributeError):
            current = None
        if current is not None:
            ws.cell(row=new_row, column=new_col).value = current
            csv_rows.append([today.strftime("%Y-%m-%d"), unit, current])
        new_listings.append((unit, current, site))
        active_rows.append(new_row)

    # Append to CSV price history (git-diffable, append-only)
    is_new_csv = not CSV_PATH.exists()
    with open(CSV_PATH, "a", newline="") as f:
        writer = csv.writer(f)
        if is_new_csv:
            writer.writerow(["date", "unit", "total_price"])
        writer.writerows(csv_rows)

    return {
        "price_changes": price_changes,
        "new_listings": new_listings,
        "off_market": off_market,
        "spec_flags": spec_flags,
        "unchanged": unchanged,
    }


def recalc_workbook():
    """Recalculate cached formula values via headless LibreOffice, if available."""
    try:
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir",
             str(XLSX_PATH.parent), str(XLSX_PATH)],
            check=True, timeout=90, capture_output=True,
        )
    except Exception as e:
        print(f"::warning::Recalc skipped ({e}). Formulas are still correct, "
              f"they just need Excel/LibreOffice to open once to refresh cached values.")


def build_html(report, as_of):
    active = [u for u in report["price_changes"]] + [u for u in report["unchanged"]]
    n_drops = sum(1 for _, old, new in report["price_changes"] if old is not None and new < old)
    n_new = len(report["new_listings"])
    n_active = len(report["price_changes"]) + len(report["unchanged"])

    rows_html = ""
    for unit, old, new in report["price_changes"]:
        if old is None:
            change_txt, direction = "new", "flat"
        else:
            delta = new - old
            pct = (delta / old * 100) if old else 0
            direction = "drop" if delta < 0 else ("rise" if delta > 0 else "flat")
            arrow = "\u25bc" if delta < 0 else ("\u25b2" if delta > 0 else "\u2014")
            change_txt = f"{arrow} ${abs(delta):,.0f} ({pct:+.1f}%)"
        rows_html += (
            f'<div class="card"><div class="card-body"><div class="top">'
            f'<span class="unit">#{unit}</span></div>'
            f'<div class="price">${new:,.0f}</div>'
            f'<div class="change {direction}">{change_txt}</div></div></div>'
        )
    for unit in report["unchanged"]:
        rows_html += (
            f'<div class="card"><div class="card-body"><div class="top">'
            f'<span class="unit">#{unit}</span></div>'
            f'<div class="change flat">no change</div></div></div>'
        )

    archive_html = "".join(
        f'<div class="archive-row"><span>#{unit}</span>'
        f'<span class="p">{f"${price:,.0f}" if price else "\u2014"}</span></div>'
        for unit, price in report["off_market"]
    )

    css = """:root{--cream:#f6f3ea;--ink:#22281f;--moss-dim:#7c8f7a;--line:#dcd6c4;
--card:#fffefa;--drop:#2f6b4f;--rise:#a8472f;--flat:#9b9685}
*{box-sizing:border-box}body{margin:0;background:var(--cream);color:var(--ink);
font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;line-height:1.4}
.wrap{max-width:980px;margin:0 auto;padding:32px 20px 80px}
header h1{font-family:Georgia,serif;font-size:34px;margin:0 0 2px}
header p{margin:0;color:var(--moss-dim);font-size:13px}
.stats{display:flex;gap:18px;flex-wrap:wrap;margin:22px 0 26px}
.stat{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:10px 16px;min-width:120px}
.stat .n{font-family:ui-monospace,Menlo,monospace;font-size:20px;font-weight:600}
.stat .l{font-size:11px;color:var(--moss-dim);text-transform:uppercase;letter-spacing:.04em}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:14px}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px}
.card-body{padding:14px 16px 16px}
.top{display:flex;justify-content:space-between;align-items:baseline}
.unit{font-weight:700;font-size:16px}
.price{font-family:ui-monospace,Menlo,monospace;font-size:24px;font-weight:600}
.change{font-size:12.5px;font-weight:600;margin-top:2px}
.drop{color:var(--drop)}.rise{color:var(--rise)}.flat{color:var(--flat)}
h2.section{font-family:Georgia,serif;font-size:18px;color:var(--moss-dim);margin:34px 0 12px;font-weight:normal}
.archive-row{display:flex;justify-content:space-between;background:var(--card);
border:1px solid var(--line);border-radius:10px;padding:10px 14px;margin-bottom:8px;font-size:13.5px}
.p{font-family:ui-monospace,Menlo,monospace;color:var(--moss-dim)}"""

    html = (
        '<!doctype html><html lang="en"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        '<title>Creekside Watch</title><style>' + css + '</style></head><body><div class="wrap">'
        '<header><h1>Creekside Watch</h1>'
        f'<p>Thornton Place \u00b7 Creekside = any unit not in Building 2 or 3 \u00b7 Data as of {as_of}</p></header>'
        '<div class="stats">'
        f'<div class="stat"><div class="n">{n_active}</div><div class="l">Active units</div></div>'
        f'<div class="stat"><div class="n">{n_drops}</div><div class="l">Price drops today</div></div>'
        f'<div class="stat"><div class="n">{n_new}</div><div class="l">New today</div></div></div>'
        f'<div class="grid">{rows_html}</div>'
        + (f'<h2 class="section">Off market</h2>{archive_html}' if archive_html else '')
        + '</div></body></html>'
    )
    HTML_PATH.parent.mkdir(parents=True, exist_ok=True)
    HTML_PATH.write_text(html)


def main():
    scraped = scrape_units()
    if not scraped:
        print("No units scraped (site may have blocked headless traffic, or genuinely nothing "
              "available). Treating as a no-op - nothing written.")
        return

    wb = load_tracker()
    report = update_tracker_and_history(wb, scraped)
    wb.save(XLSX_PATH)
    recalc_workbook()

    as_of = datetime.now().strftime("%b %d, %Y at %I:%M %p UTC")
    build_html(report, as_of)

    print(f"Price changes: {len(report['price_changes'])}")
    print(f"New listings: {len(report['new_listings'])}")
    print(f"Off market: {len(report['off_market'])}")
    print(f"Spec mismatches flagged: {len(report['spec_flags'])}")
    for unit, field, tracker_val, site_val in report["spec_flags"]:
        print(f"  ::warning:: #{unit} {field} mismatch - tracker={tracker_val} site={site_val}")


if __name__ == "__main__":
    main()
