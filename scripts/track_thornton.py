#!/usr/bin/env python3
"""
Daily Thornton Place / Creekside price tracker.

Ported from Jonas's Cowork routine (thornton-track-cloud) to run as a
GitHub Actions job with no cloud storage dependency - everything lives
in this repo. Reads/writes:

  data/Thornton_Tracker.xlsx   canonical workbook (must already exist -
                                seed it with your current tracker before
                                the first run; this script never creates
                                one from scratch, to protect price history)
  data/price-history.csv       flat, git-diffable append log (one row per
                                unit per run where the price is known)
  docs/index.html               self-contained dashboard, served by GitHub Pages -
                                each unit card includes a price-history sparkline
                                built from price-history.csv

Constraints preserved from the original routine:
  - Creekside = any unit NOT in Building 2 or 3 (defined by exclusion)
  - Uses TOTAL monthly price, not base rent
  - Never overwrites spec columns (beds/bath/sqft/layout) - flags mismatches only
  - A failed/empty scrape is treated as a no-op, never as "everything off market"
"""

import csv
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "Thornton_Tracker.xlsx"
CSV_PATH = ROOT / "data" / "price-history.csv"
HTML_PATH = ROOT / "docs" / "index.html"

SITE_URL = "https://thornton-place.com/floorplans/"


def scrape_units():
    """Headless-Chromium scrape of the SightMap floor widget. Returns a
    list of dicts, or [] if scraping failed / found nothing (caller must
    treat that as a no-op, not as 'everything went off market')."""
    from playwright.sync_api import sync_playwright

    js = r"""
    (async () => {
      const sleep = ms => new Promise(r => setTimeout(r, ms));
      const all = new Map();
      const floorItems = Array.from(
        document.querySelectorAll('.jd-fp-map-embed__floors-item.swiper-slide')
      );
      const floorsWithAvail = floorItems.filter(el => el.className.includes('--has-avail'));
      const extract = () => {
        document.querySelectorAll('.jd-fp-unit-card.jd-fp-unit-card--row').forEach(c => {
          const txt = c.textContent.trim().replace(/\s+/g, ' ');
          const unitM  = txt.match(/#(\d-\d{3}[A-Z]?)/);
          const totalM = txt.match(/\$([\d,]+\.\d{2})\s*\/mo/);
          const baseM  = txt.match(/\$([\d,]+)\s*Base Rent/);
          const availM = txt.match(/Available\s+(Now|[A-Z][a-z]+\s+\d+)/);
          const bedsM  = txt.match(/(Studio|\d+)\s*bed/i);
          const bathM  = txt.match(/([\d.]+)\s*bath/i);
          const sqftM  = txt.match(/([\d,]+)\s*sq\.?\s*ft/i);
          const layoutM = txt.match(/^([A-Z]\d+[A-Z]?)\s/);
          if (unitM && !all.has(unitM[1])) {
            all.set(unitM[1], { unit: unitM[1], total: totalM?.[1], base: baseM?.[1],
              avail: availM?.[1], beds: bedsM?.[1], bath: bathM?.[1],
              sqft: sqftM?.[1], layout: layoutM?.[1], raw: txt });
          }
        });
      };
      extract();
      for (const item of floorsWithAvail) { item.click(); await sleep(1200); extract(); }
      return Array.from(all.values());
    })()
    """

    debug_dir = ROOT / "data" / "debug"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--use-gl=swiftshader",  # software WebGL - the SightMap widget needs a GL context
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                ],
            )
            page = browser.new_page(
                viewport={"width": 1440, "height": 900},
                user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                             "AppleWebKit/537.36 (KHTML, like Gecko) "
                             "Chrome/125.0.0.0 Safari/537.36"),
            )
            # Hide the most common automation fingerprint
            page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

            page.goto(SITE_URL, wait_until="networkidle", timeout=45000)
            page.wait_for_timeout(3500)
            units = page.evaluate(js)

            if not units:
                # Leave evidence behind so a human (or Claude) can see what actually rendered
                debug_dir.mkdir(parents=True, exist_ok=True)
                page.screenshot(path=str(debug_dir / "last_failure.png"), full_page=True)
                (debug_dir / "last_failure.html").write_text(page.content())
                print("::warning::Zero units extracted - saved data/debug/last_failure.png "
                      "and last_failure.html for inspection.")

            browser.close()
    except Exception as e:
        print(f"::warning::Scrape failed: {e}")
        return []

    # Creekside = NOT building 2 or 3
    creekside = [u for u in units if u.get("unit") and not u["unit"].startswith(("2-", "3-"))]
    return creekside


def load_tracker():
    if not XLSX_PATH.exists():
        print(f"::error::{XLSX_PATH} not found. Upload your existing Thornton Tracker.xlsx "
              f"into data/ before the first run - this script will not create one from scratch.")
        sys.exit(1)
    return load_workbook(XLSX_PATH)


def find_active_rows(ws):
    """Overview sheet: A=Unit#, B=Creekside?, ... row with A='— Off Market —'
    separates active rows from the off-market archive."""
    active_rows = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        if isinstance(val, str) and "Off Market" in val:
            break
        active_rows.append(row)
    return active_rows


def last_recorded_price(ws, row, last_date_col):
    """Scan newest -> left through the date columns for the most recent numeric price."""
    for col in range(last_date_col, 8, -1):  # column I=9 is the Feb baseline; date cols start at J=10
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            return v, col
    return None, None


def compose_specs(site):
    """Build a 'Studio, 1 bath' / '2 bed, 1.5 bath' string for a freshly
    scraped unit, matching the format already used in the Specs column."""
    beds = site.get("beds")
    bath = site.get("bath")
    parts = []
    if beds:
        parts.append("Studio" if str(beds).lower() == "studio" else f"{beds} bed")
    if bath:
        parts.append(f"{bath} bath")
    return ", ".join(parts)


def read_row_specs(ws, row):
    return {
        "location": ws.cell(row=row, column=3).value,
        "specs": ws.cell(row=row, column=4).value,
        "sqft": ws.cell(row=row, column=5).value,
        "layout": ws.cell(row=row, column=6).value,
        "avail": ws.cell(row=row, column=7).value,
    }


def update_tracker_and_history(wb, scraped):
    ws = wb.active  # Overview sheet
    today = datetime.now()
    active_rows = find_active_rows(ws)

    # Find the current last date column (rightmost column before "$ Change")
    header_row = 1
    last_date_col = None
    change_col = None
    for col in range(ws.max_column, 0, -1):
        header = ws.cell(row=header_row, column=col).value
        if header == "$ Change":
            change_col = col
        if isinstance(header, datetime) and last_date_col is None:
            last_date_col = col
    if last_date_col is None or change_col is None:
        print("::error::Could not locate date columns / '$ Change' header - tracker layout unexpected.")
        sys.exit(1)

    tracker_by_unit = {}
    for row in active_rows:
        unit = ws.cell(row=row, column=1).value
        if unit:
            tracker_by_unit[str(unit).lstrip("#")] = row

    scraped_by_unit = {u["unit"]: u for u in scraped}

    price_changes, new_listings, off_market, spec_flags, unchanged = [], [], [], [], []
    active_units = []  # full display info (specs/sqft/price/etc.) for the dashboard

    # Insert a new date column before "$ Change"
    new_col = change_col
    ws.insert_cols(new_col)
    change_col += 1
    # copy style from previous date column header
    src_header_cell = ws.cell(row=header_row, column=last_date_col)
    new_header_cell = ws.cell(row=header_row, column=new_col)
    new_header_cell.value = today
    new_header_cell.number_format = src_header_cell.number_format
    if src_header_cell.has_style:
        new_header_cell.font = src_header_cell.font.copy()
        new_header_cell.fill = src_header_cell.fill.copy()
        new_header_cell.alignment = src_header_cell.alignment.copy()
        new_header_cell.border = src_header_cell.border.copy()

    csv_rows = []

    for unit, row in tracker_by_unit.items():
        last_price, _ = last_recorded_price(ws, row, last_date_col)
        site = scraped_by_unit.get(unit)
        specs = read_row_specs(ws, row)

        if site is None:
            ws.cell(row=row, column=8).value = "Off Market"  # Notes column H
            off_market.append((unit, last_price, specs))
            continue

        try:
            current = float(site["total"].replace(",", "")) if site.get("total") else None
        except (ValueError, AttributeError):
            current = None

        if current is not None:
            if last_price is None or abs(current - last_price) > 0.005:
                ws.cell(row=row, column=new_col).value = current
                price_changes.append((unit, last_price, current))
                if last_price is None:
                    status = "flat"  # first price ever recorded - nothing to compare against
                else:
                    status = "drop" if current < last_price else "rise"
            else:
                unchanged.append(unit)
                status = "flat"
            csv_rows.append([today.strftime("%Y-%m-%d"), unit, current])
            active_units.append({"unit": unit, "price": current, "old_price": last_price,
                                  "status": status, **specs})

        # spec sanity check only - never auto-write D/E/F
        beds = ws.cell(row=row, column=4).value
        sqft = ws.cell(row=row, column=5).value
        if site.get("sqft") and sqft and str(site["sqft"]).replace(",", "") != str(sqft).replace(",", ""):
            spec_flags.append((unit, "sqft", sqft, site["sqft"]))

    # New listings not already in tracker
    for unit, site in scraped_by_unit.items():
        if unit in tracker_by_unit:
            continue
        new_row = max(active_rows) + 1 if active_rows else 2
        ws.insert_rows(new_row)
        ws.cell(row=new_row, column=1).value = f"#{unit}"
        ws.cell(row=new_row, column=2).value = "\u2705"
        specs_text = compose_specs(site)
        ws.cell(row=new_row, column=4).value = specs_text
        ws.cell(row=new_row, column=5).value = site.get("sqft", "")
        ws.cell(row=new_row, column=6).value = site.get("layout", "")
        ws.cell(row=new_row, column=7).value = site.get("avail", "")
        try:
            current = float(site["total"].replace(",", "")) if site.get("total") else None
        except (ValueError, AttributeError):
            current = None
        if current is not None:
            ws.cell(row=new_row, column=new_col).value = current
            csv_rows.append([today.strftime("%Y-%m-%d"), unit, current])
        new_listings.append((unit, current, site))
        active_rows.append(new_row)
        active_units.append({
            "unit": unit, "price": current, "old_price": None, "status": "new",
            "location": None, "specs": specs_text, "sqft": site.get("sqft"),
            "layout": site.get("layout"), "avail": site.get("avail"),
        })

    # Append to CSV price history (git-diffable, append-only)
    is_new_csv = not CSV_PATH.exists()
    with open(CSV_PATH, "a", newline="") as f:
        writer = csv.writer(f)
        if is_new_csv:
            writer.writerow(["date", "unit", "total_price"])
        writer.writerows(csv_rows)

    return {
        "price_changes": price_changes,
        "new_listings": new_listings,
        "off_market": off_market,
        "spec_flags": spec_flags,
        "unchanged": unchanged,
        "active_units": active_units,
    }


def recalc_workbook():
    """Recalculate cached formula values via headless LibreOffice, if available."""
    try:
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir",
             str(XLSX_PATH.parent), str(XLSX_PATH)],
            check=True, timeout=90, capture_output=True,
        )
    except Exception as e:
        print(f"::warning::Recalc skipped ({e}). Formulas are still correct, "
              f"they just need Excel/LibreOffice to open once to refresh cached values.")


STATUS_PRIORITY = {"new": 0, "drop": 1, "rise": 2, "flat": 3}
ARROW_DOWN = "\u25bc"
ARROW_UP = "\u25b2"
EM_DASH = "\u2014"
MIDDOT = "\u00b7"

BASE_CSS = """:root{--cream:#f6f3ea;--ink:#22281f;--moss-dim:#7c8f7a;--line:#dcd6c4;
--card:#fffefa;--drop:#2f6b4f;--rise:#a8472f;--flat:#9b9685;--accent:#7c8f7a}
@media (prefers-color-scheme:dark){:root{--cream:#1b1d18;--ink:#eae7da;--moss-dim:#9aa38f;
--line:#33362c;--card:#242720;--drop:#5fbf8f;--rise:#e08668;--flat:#7c8069;--accent:#5fbf8f}}
*{box-sizing:border-box}body{margin:0;background:var(--cream);color:var(--ink);
font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;line-height:1.4}
.wrap{max-width:1040px;margin:0 auto;padding:32px 20px 80px}
header h1{font-family:Georgia,serif;font-size:34px;margin:0 0 2px}
header p{margin:0;color:var(--moss-dim);font-size:13px}
.stats{display:flex;gap:18px;flex-wrap:wrap;margin:22px 0 26px}
.stat{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:10px 16px;min-width:120px}
.stat .n{font-family:ui-monospace,Menlo,monospace;font-size:20px;font-weight:600}
.stat .l{font-size:11px;color:var(--moss-dim);text-transform:uppercase;letter-spacing:.04em}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(270px,1fr));gap:14px}
.card{background:var(--card);border:1px solid var(--line);border-left:3px solid var(--line);border-radius:12px}
.card-new{border-left-color:var(--accent)}.card-drop{border-left-color:var(--drop)}
.card-rise{border-left-color:var(--rise)}
.card-body{padding:14px 16px 16px}
.top{display:flex;justify-content:space-between;align-items:center;gap:8px}
.unit{font-weight:700;font-size:16px}
.layout{margin-left:auto;font-size:11px;color:var(--moss-dim);border:1px solid var(--line);
border-radius:6px;padding:1px 6px}
.badge-new{font-size:10px;font-weight:700;letter-spacing:.04em;color:var(--card);
background:var(--accent);border-radius:6px;padding:2px 6px}
.specs{font-size:13px;color:var(--moss-dim);margin-top:6px}
.location{font-size:12px;color:var(--moss-dim);margin-top:2px}
.avail{font-size:12px;color:var(--moss-dim);margin-top:2px}
.price{font-family:ui-monospace,Menlo,monospace;font-size:24px;font-weight:600;margin-top:10px}
.price-na{color:var(--moss-dim);font-size:18px}
.persqft{font-size:11.5px;color:var(--moss-dim);margin-top:1px}
.change{font-size:12.5px;font-weight:600;margin-top:6px}
.drop{color:var(--drop)}.rise{color:var(--rise)}.flat{color:var(--flat)}.new{color:var(--accent)}
h2.section{font-family:Georgia,serif;font-size:18px;color:var(--moss-dim);margin:34px 0 12px;font-weight:normal}
.archive-row{display:flex;justify-content:space-between;align-items:center;gap:12px;background:var(--card);
border:1px solid var(--line);border-radius:10px;padding:10px 14px;margin-bottom:8px;font-size:13.5px}
.au{flex:0 0 auto;font-weight:600}
.archive-specs{flex:1 1 auto;color:var(--moss-dim);font-size:12.5px;text-align:right;margin-right:8px}
.p{font-family:ui-monospace,Menlo,monospace;color:var(--moss-dim);flex:0 0 auto}
.empty{color:var(--moss-dim);font-size:13.5px;padding:20px 0}
.spark-wrap{margin-top:10px}
.sparkline{width:100%;height:auto;display:block}
.spark-line{stroke:var(--moss-dim);stroke-width:2;stroke-linecap:round;stroke-linejoin:round;opacity:.6}
.spark-dot-hist{fill:var(--card);stroke:var(--moss-dim);stroke-width:1.5;opacity:.6}
.spark-dot-drop{fill:var(--drop);stroke:var(--card);stroke-width:2}
.spark-dot-rise{fill:var(--rise);stroke:var(--card);stroke-width:2}
.spark-dot-flat{fill:var(--flat);stroke:var(--card);stroke-width:2}
.spark-dot-new{fill:var(--accent);stroke:var(--card);stroke-width:2}
.spark-hit{fill:transparent;cursor:default}
.spark-hit:hover,.spark-hit:focus{fill:var(--ink);opacity:.08;outline:none}
.no-history{color:var(--moss-dim);font-size:12px;font-style:italic;margin-top:10px}
.chart-tooltip{position:fixed;transform:translate(-50%,-100%);background:var(--ink);color:var(--cream);
font-size:12px;font-family:ui-monospace,Menlo,monospace;font-weight:600;padding:5px 9px;border-radius:6px;
pointer-events:none;opacity:0;transition:opacity .1s ease;z-index:1000;white-space:nowrap}
.chart-tooltip.visible{opacity:1}"""

TOOLTIP_JS = """
(function(){
  var tip = document.getElementById('chart-tooltip');
  if (!tip) return;
  function show(el){
    var date = el.getAttribute('data-date'), price = el.getAttribute('data-price');
    if (!date || !price) return;
    tip.textContent = date + ': ' + price;
    var r = el.getBoundingClientRect();
    tip.style.left = (r.left + r.width / 2) + 'px';
    tip.style.top = (r.top - 6) + 'px';
    tip.classList.add('visible');
  }
  function hide(){ tip.classList.remove('visible'); }
  document.addEventListener('pointerover', function(e){
    var el = e.target.closest && e.target.closest('.spark-hit');
    if (el) show(el);
  });
  document.addEventListener('pointerout', function(e){
    var el = e.target.closest && e.target.closest('.spark-hit');
    if (el) hide();
  });
  document.addEventListener('focusin', function(e){
    var el = e.target.closest && e.target.closest('.spark-hit');
    if (el) show(el);
  });
  document.addEventListener('focusout', function(e){
    var el = e.target.closest && e.target.closest('.spark-hit');
    if (el) hide();
  });
})();
"""


def read_price_history():
    """Parses data/price-history.csv into {unit: [(date, price), ...]} sorted
    chronologically. Returns {} if the file doesn't exist yet (first-ever run)."""
    by_unit = {}
    if not CSV_PATH.exists():
        return by_unit
    with open(CSV_PATH, newline="") as f:
        for row in csv.DictReader(f):
            try:
                price = float(row["total_price"])
            except (KeyError, ValueError):
                continue
            by_unit.setdefault(row["unit"], []).append((row["date"], price))
    for series in by_unit.values():
        series.sort(key=lambda p: p[0])
    return by_unit


def render_card(u, series):
    """One dashboard tile for an active unit: unit #, layout, specs, sqft,
    availability, price, $/sqft, today's change vs. yesterday, and a price
    history sparkline (series: [(date, price), ...] for this unit, oldest first)."""
    layout_badge = f'<span class="layout">{u["layout"]}</span>' if u.get("layout") else ""
    badge = '<span class="badge-new">NEW</span>' if u["status"] == "new" else ""

    spec_bits = []
    if u.get("specs"):
        spec_bits.append(str(u["specs"]))
    if u.get("sqft"):
        spec_bits.append(f"{u['sqft']} sq ft")
    specs_line = f' {MIDDOT} '.join(spec_bits)
    location_line = f'<div class="location">{u["location"]}</div>' if u.get("location") else ""
    avail_line = ""
    if u.get("avail"):
        avail_text = str(u["avail"])
        prefix = "" if avail_text.lower().startswith("available") else "Available "
        avail_line = f'<div class="avail">{prefix}{avail_text}</div>'

    price = u.get("price")
    if price is None:
        price_html = f'<div class="price price-na">{EM_DASH}</div>'
        per_sqft_html = ""
    else:
        price_html = f'<div class="price">${price:,.0f}</div>'
        per_sqft_html = ""
        if u.get("sqft"):
            try:
                per_sqft = price / float(str(u["sqft"]).replace(",", ""))
                per_sqft_html = f'<div class="persqft">${per_sqft:.2f}/sq ft</div>'
            except (ValueError, ZeroDivisionError):
                pass

    status = u["status"]
    old = u.get("old_price")
    if status == "new":
        change_txt = "new listing"
    elif status == "flat":
        change_txt = "no change"
    else:
        delta = price - old
        pct = (delta / old * 100) if old else 0
        arrow = ARROW_DOWN if delta < 0 else ARROW_UP
        change_txt = f"{arrow} ${abs(delta):,.0f} ({pct:+.1f}%)"

    return (
        f'<div class="card card-{status}"><div class="card-body">'
        f'<div class="top"><span class="unit">#{u["unit"]}</span>{layout_badge}{badge}</div>'
        f'<div class="specs">{specs_line}</div>'
        f'{location_line}{avail_line}'
        f'{price_html}{per_sqft_html}'
        f'<div class="change {status}">{change_txt}</div>'
        f'<div class="spark-wrap">{render_sparkline(series, status=status)}</div>'
        '</div></div>'
    )


def render_archive_row(unit, price, specs):
    price_txt = f"${price:,.0f}" if price else EM_DASH
    spec_bits = []
    if specs.get("specs"):
        spec_bits.append(str(specs["specs"]))
    if specs.get("sqft"):
        spec_bits.append(f"{specs['sqft']} sq ft")
    spec_txt = f' {MIDDOT} '.join(spec_bits)
    return (
        f'<div class="archive-row"><span class="au">#{unit}</span>'
        f'<span class="archive-specs">{spec_txt}</span>'
        f'<span class="p">{price_txt}</span></div>'
    )


def build_html(report, as_of):
    active_units = sorted(
        report["active_units"],
        key=lambda u: (STATUS_PRIORITY.get(u["status"], 9), u["unit"]),
    )
    n_drops = sum(1 for u in active_units if u["status"] == "drop")
    n_new = len(report["new_listings"])
    n_active = len(active_units)
    priced = [u["price"] for u in active_units if u.get("price") is not None]
    avg_price = sum(priced) / len(priced) if priced else None

    price_history = read_price_history()
    rows_html = "".join(render_card(u, price_history.get(u["unit"], [])) for u in active_units)
    archive_html = "".join(render_archive_row(unit, price, specs) for unit, price, specs in report["off_market"])

    avg_stat = (
        f'<div class="stat"><div class="n">${avg_price:,.0f}</div><div class="l">Avg. price</div></div>'
        if avg_price is not None else ""
    )

    html = (
        '<!doctype html><html lang="en"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        '<link rel="icon" href="data:,%F0%9F%8F%A2">'
        '<title>Creekside Watch</title><style>' + BASE_CSS + '</style></head><body><div class="wrap">'
        '<header><h1>Creekside Watch</h1>'
        f'<p>Thornton Place {MIDDOT} Creekside = any unit not in Building 2 or 3 {MIDDOT} Data as of {as_of}</p></header>'
        '<div class="stats">'
        f'<div class="stat"><div class="n">{n_active}</div><div class="l">Active units</div></div>'
        f'<div class="stat"><div class="n">{n_drops}</div><div class="l">Price drops today</div></div>'
        f'<div class="stat"><div class="n">{n_new}</div><div class="l">New today</div></div>'
        f'{avg_stat}</div>'
        + (f'<div class="grid">{rows_html}</div>' if rows_html else '<div class="empty">No active Creekside units right now.</div>')
        + (f'<h2 class="section">Off market</h2>{archive_html}' if archive_html else '')
        + '</div><div id="chart-tooltip" class="chart-tooltip" role="tooltip"></div>'
        + '<script>' + TOOLTIP_JS + '</script></body></html>'
    )
    HTML_PATH.parent.mkdir(parents=True, exist_ok=True)
    HTML_PATH.write_text(html)


def _fmt_axis_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%b %-d")
    except ValueError:
        return date_str


def _svg_points(series, width, height, pad_l=4, pad_r=4, pad_t=10, pad_b=10):
    """series: [(label, value), ...]. Single point is right-aligned so it sits
    where a multi-point line's endpoint would; flat/constant series draw as a
    flat line rather than dividing by zero."""
    n = len(series)
    if n == 0:
        return []
    values = [v for _, v in series]
    lo, hi = min(values), max(values)
    if hi - lo < 1e-9:
        lo, hi = lo - 1, hi + 1
    usable_w = width - pad_l - pad_r
    usable_h = height - pad_t - pad_b
    points = []
    for i, (_, v) in enumerate(series):
        x = pad_l + (usable_w if n == 1 else usable_w * i / (n - 1))
        y = pad_t + usable_h * (1 - (v - lo) / (hi - lo))
        points.append((x, y))
    return points


def render_sparkline(series, width=220, height=54, status="flat"):
    """Compact per-unit trend: muted line, status-colored endpoint dot. Each
    point gets an oversized transparent hit-circle on top (the visible dot is
    only 5-8px, far too small to hover reliably) carrying data-date/data-price
    attributes - a page-level script (see build_html) reads these and drives
    a real tooltip, since native SVG <title> tooltips are unreliable across
    browsers/trackpads."""
    if not series:
        return '<div class="no-history">No price history yet</div>'
    points = _svg_points(series, width, height)
    marks = []
    if len(points) > 1:
        path_d = "M " + " L ".join(f"{x:.1f},{y:.1f}" for x, y in points)
        marks.append(f'<path d="{path_d}" class="spark-line" fill="none"/>')
    hit_areas = []
    for i, ((date, price), (x, y)) in enumerate(zip(series, points)):
        is_last = i == len(series) - 1
        cls = f"spark-dot-{status}" if is_last else "spark-dot-hist"
        r = 4 if is_last else 2.5
        marks.append(f'<circle class="spark-dot {cls}" cx="{x:.1f}" cy="{y:.1f}" r="{r}"/>')
        hit_areas.append(
            f'<circle class="spark-hit" cx="{x:.1f}" cy="{y:.1f}" r="12" tabindex="0" '
            f'data-date="{_fmt_axis_date(date)}" data-price="${price:,.0f}"/>'
        )
    marks.extend(hit_areas)  # hit-circles last so they sit on top and actually receive hover
    return (
        f'<svg class="sparkline" viewBox="0 0 {width} {height}" role="img" '
        f'aria-label="Price trend, {len(series)} data point(s)">{"".join(marks)}</svg>'
    )


def main():
    scraped = scrape_units()
    if not scraped:
        print("No units scraped (site may have blocked headless traffic, or genuinely nothing "
              "available). Treating as a no-op - nothing written.")
        return

    wb = load_tracker()
    report = update_tracker_and_history(wb, scraped)
    wb.save(XLSX_PATH)
    recalc_workbook()

    as_of = datetime.now().strftime("%b %d, %Y at %I:%M %p UTC")
    build_html(report, as_of)

    print(f"Price changes: {len(report['price_changes'])}")
    print(f"New listings: {len(report['new_listings'])}")
    print(f"Off market: {len(report['off_market'])}")
    print(f"Spec mismatches flagged: {len(report['spec_flags'])}")
    for unit, field, tracker_val, site_val in report["spec_flags"]:
        print(f"  ::warning:: #{unit} {field} mismatch - tracker={tracker_val} site={site_val}")


if __name__ == "__main__":
    main()
